from io import BytesIO

from typing import (
    Any,
    Dict,
    List,
)

from openpyxl import (
    load_workbook,
)


BUSINESS_SHEETS = [
    "users",
    "profiles",
    "skills",
    "user_skills",
    "skill_evidence",
    "skill_verifications",
    "resume_files",
    "opportunities",
    "opportunity_skills",
    "applications",
    "mentorships",
    "mentorship_sessions",
    "challenges",
    "challenge_skills",
    "challenge_submissions",
    "notifications",
]


KEY_COLUMNS = {
    "users":
        "user_key",

    "profiles":
        "user_key",

    "skills":
        "skill_key",

    "user_skills":
        "user_skill_key",

    "skill_evidence":
        "evidence_key",

    "skill_verifications":
        "verification_key",

    "resume_files":
        "resume_key",

    "opportunities":
        "opportunity_key",

    "opportunity_skills":
        "opportunity_skill_key",

    "applications":
        "application_key",

    "mentorships":
        "mentorship_key",

    "mentorship_sessions":
        "session_key",

    "challenges":
        "challenge_key",

    "challenge_skills":
        "challenge_skill_key",

    "challenge_submissions":
        "submission_key",

    "notifications":
        "notification_key",
}


REFERENCE_RULES = [
    (
        "profiles",
        "user_key",
        "users",
    ),

    (
        "user_skills",
        "user_key",
        "users",
    ),
    (
        "user_skills",
        "skill_key",
        "skills",
    ),

    (
        "skill_evidence",
        "user_skill_key",
        "user_skills",
    ),

    (
        "skill_verifications",
        "evidence_key",
        "skill_evidence",
    ),
    (
        "skill_verifications",
        "verifier_user_key",
        "users",
    ),

    (
        "resume_files",
        "user_key",
        "users",
    ),

    (
        "opportunities",
        "employer_user_key",
        "users",
    ),

    (
        "opportunity_skills",
        "opportunity_key",
        "opportunities",
    ),
    (
        "opportunity_skills",
        "skill_key",
        "skills",
    ),

    (
        "applications",
        "opportunity_key",
        "opportunities",
    ),
    (
        "applications",
        "student_user_key",
        "users",
    ),
    (
        "applications",
        "resume_key",
        "resume_files",
    ),

    (
        "mentorships",
        "student_user_key",
        "users",
    ),
    (
        "mentorships",
        "mentor_user_key",
        "users",
    ),

    (
        "mentorship_sessions",
        "mentorship_key",
        "mentorships",
    ),

    (
        "challenges",
        "employer_user_key",
        "users",
    ),

    (
        "challenge_skills",
        "challenge_key",
        "challenges",
    ),
    (
        "challenge_skills",
        "skill_key",
        "skills",
    ),

    (
        "challenge_submissions",
        "challenge_key",
        "challenges",
    ),
    (
        "challenge_submissions",
        "student_user_key",
        "users",
    ),

    (
        "notifications",
        "user_key",
        "users",
    ),
]


def read_sample_workbook(
    file_bytes: bytes,
) -> Dict[
    str,
    List[
        Dict[str, Any]
    ],
]:

    workbook = load_workbook(
        filename=BytesIO(
            file_bytes
        ),
        read_only=True,
        data_only=True,
    )

    missing = [
        sheet
        for sheet in BUSINESS_SHEETS
        if sheet not in workbook.sheetnames
    ]

    if missing:
        raise ValueError(
            "Missing required sheets: "
            + ", ".join(
                missing
            )
        )

    result = {}

    for sheet_name in (
        BUSINESS_SHEETS
    ):

        worksheet = workbook[
            sheet_name
        ]

        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            first_row = next(rows)

        except StopIteration:
            result[
                sheet_name
            ] = []

            continue

        headers = [
            (
                str(value).strip()
                if value is not None
                else ""
            )
            for value
            in first_row
        ]

        sheet_rows = []

        for (
            row_number,
            row,
        ) in enumerate(
            rows,
            start=2,
        ):

            if all(
                value in (
                    None,
                    "",
                )
                for value in row
            ):
                continue

            item = {}

            for (
                index,
                value,
            ) in enumerate(
                row
            ):

                if (
                    index
                    >= len(headers)
                ):
                    continue

                header = (
                    headers[index]
                )

                if not header:
                    continue

                item[
                    header
                ] = value

            item[
                "_row_number"
            ] = row_number

            sheet_rows.append(
                item
            )

        result[
            sheet_name
        ] = sheet_rows

    return result


def validate_sample_workbook(
    data,
):

    errors = []

    lookup = {}

    for (
        sheet_name,
        key_column,
    ) in KEY_COLUMNS.items():

        seen = set()

        lookup[
            sheet_name
        ] = set()

        for row in data.get(
            sheet_name,
            [],
        ):

            value = row.get(
                key_column
            )

            if value in (
                None,
                "",
            ):
                errors.append(
                    {
                        "sheet":
                            sheet_name,

                        "row":
                            row.get(
                                "_row_number"
                            ),

                        "field":
                            key_column,

                        "message":
                            "Key is required.",
                    }
                )

                continue

            value = str(
                value
            ).strip()

            if value in seen:
                errors.append(
                    {
                        "sheet":
                            sheet_name,

                        "row":
                            row.get(
                                "_row_number"
                            ),

                        "field":
                            key_column,

                        "message":
                            (
                                "Duplicate key: "
                                f"{value}"
                            ),
                    }
                )

            seen.add(value)

            lookup[
                sheet_name
            ].add(value)

    for (
        source_sheet,
        source_field,
        target_sheet,
    ) in REFERENCE_RULES:

        target_values = (
            lookup.get(
                target_sheet,
                set(),
            )
        )

        for row in data.get(
            source_sheet,
            [],
        ):

            reference = row.get(
                source_field
            )

            if reference in (
                None,
                "",
            ):
                continue

            reference = str(
                reference
            ).strip()

            if (
                reference
                not in target_values
            ):

                errors.append(
                    {
                        "sheet":
                            source_sheet,

                        "row":
                            row.get(
                                "_row_number"
                            ),

                        "field":
                            source_field,

                        "message":
                            (
                                "Unknown reference: "
                                f"{reference}"
                            ),
                    }
                )

    roles = {
        str(
            row.get(
                "user_key"
            )
        ):
            str(
                row.get(
                    "role",
                    "",
                )
            ).lower()

        for row
        in data.get(
            "users",
            [],
        )
    }

    allowed_roles = {
        "student",
        "employer",
        "mentor",
    }

    for row in data.get(
        "users",
        [],
    ):

        role = str(
            row.get(
                "role",
                "",
            )
        ).lower()

        if (
            role
            not in allowed_roles
        ):

            errors.append(
                {
                    "sheet":
                        "users",

                    "row":
                        row.get(
                            "_row_number"
                        ),

                    "field":
                        "role",

                    "message":
                        (
                            "Role must be student, "
                            "employer, or mentor."
                        ),
                }
            )

    for row in data.get(
        "opportunities",
        [],
    ):

        key = str(
            row.get(
                "employer_user_key",
                "",
            )
        )

        if (
            roles.get(key)
            != "employer"
        ):

            errors.append(
                {
                    "sheet":
                        "opportunities",

                    "row":
                        row.get(
                            "_row_number"
                        ),

                    "field":
                        "employer_user_key",

                    "message":
                        (
                            "Opportunity must "
                            "reference an employer."
                        ),
                }
            )

    for row in data.get(
        "applications",
        [],
    ):

        key = str(
            row.get(
                "student_user_key",
                "",
            )
        )

        if (
            roles.get(key)
            != "student"
        ):

            errors.append(
                {
                    "sheet":
                        "applications",

                    "row":
                        row.get(
                            "_row_number"
                        ),

                    "field":
                        "student_user_key",

                    "message":
                        (
                            "Application must "
                            "reference a student."
                        ),
                }
            )

    return errors